import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side


class ExcelHandler:
    def __init__(self, template_file):
        self.template_file = template_file

    def duplicate_file(self, version=1):
        """Creates a duplicate of the template file with a new version number."""
        base_name = f"com_result_v{version}.xlsx"
        if os.path.exists(base_name):
            return self.duplicate_file(version + 1)
        # Load the template and save it as a new file
        df = pd.read_excel(self.template_file)
        df.to_excel(base_name, index=False)
        return base_name

    def add_new_headers(self, df):
        """Adds new headers for additional columns."""
        df['current_remote_host'] = None
        df['current_remote_int'] = None
        df['status'] = None
        df['current_TX'] = None
        df['current_RX'] = None
        df['status_TX'] = None
        df['status_RX'] = None
        return df

    def get_base_hostname(self, hostname):
        #Extracts the base hostname by removing `.X@Y.ru` if present, otherwise keeps the full hostname.
        match = re.search(r'(.+?)\.[^.]+.[^.]+\.ru$', hostname)  # Look for `.X@Y.ru` pattern
        return match.group(1) if match else hostname  # Extract part before `.X@Y.ru`, else return full hostname

    def normalize_interface_name(self, interface):
        # Convert short interface names to full names for comparison.
        return (interface
                .replace("GigabitEthernet", "Gi")
                .replace("FastEthernet", "Fa")
                .replace("Ethernet", "Eth"))

    def add_borders_to_excel(self, file_path):
        """Adds borders to all cells in the Excel file."""
        wb = load_workbook(file_path)
        sheet = wb.active

        # Define border style
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        # Iterate through all rows and columns to add borders
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border

        # Save the file with borders
        wb.save(file_path)

    def populate_and_compare(self, new_data, signal_data):
        """Handles the entire process of populating data and comparing rows."""
        result_file = self.duplicate_file()  # Duplicate file to keep original intact
        df = pd.read_excel(result_file)

        # Add new headers for additional columns
        df = self.add_new_headers(df)

        # New data to DataFrame
        new_data_df = pd.DataFrame(new_data, columns=['local_host', 'local_int', 'remote_host', 'remote_int'])
        signal_data_df = pd.DataFrame(signal_data, columns=['local_host', 'local_int', 'TX', 'RX'])

        # Populate new data
        for idx, row in new_data_df.iterrows():
            local_host, local_int, remote_host, remote_int = row
            base_remote_host = self.get_base_hostname(remote_host)

            match = df[(df['local_host'] == local_host) & (df['local_int'] == local_int) &
                       (df['remote_host'] == remote_host) & (df['remote_int'] == remote_int)]

            if not match.empty:
                df.loc[match.index, 'current_remote_host'] = remote_host
                df.loc[match.index, 'current_remote_int'] = remote_int
                df.loc[match.index, 'status'] = 'OK'
            else:
                partial_match = df[(df['local_host'] == local_host) & (df['local_int'] == local_int)]

                if not partial_match.empty:
                    for i in partial_match.index:
                        expected_base_host = self.get_base_hostname(df.at[i, 'remote_host'])
                        if (expected_base_host == base_remote_host and
                                self.normalize_interface_name(df.at[i, 'remote_int']) == self.normalize_interface_name(
                                    remote_int)):
                            df.loc[i, 'current_remote_host'] = remote_host
                            df.loc[i, 'current_remote_int'] = remote_int
                            df.loc[i, 'status'] = 'OK'
                            break
                    else:
                        df.loc[partial_match.index, 'current_remote_host'] = remote_host
                        df.loc[partial_match.index, 'current_remote_int'] = remote_int
                        df.loc[partial_match.index, 'status'] = 'Ошибка'
                else:
                    new_entry = pd.DataFrame([row], columns=['local_host', 'local_int', 'remote_host', 'remote_int'])
                    with pd.ExcelWriter(result_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        new_entry.to_excel(writer, sheet_name="Shouldn't be there", index=False)

        # Populate signal data
        for idx, signal_row in signal_data_df.iterrows():
            local_host, local_int, tx, rx = signal_row
            signal_match = df[(df['local_host'] == local_host) & (df['local_int'] == local_int)]
            if not signal_match.empty:
                df.loc[signal_match.index, 'current_TX'] = tx
                df.loc[signal_match.index, 'current_RX'] = rx

                # Evaluate TX and RX status
                def evaluate_status(power_values):
                    if power_values in ("--", "-", "-40", "-40.00") or all(v.strip() in ("--", "-", "-40", "-40.00") for v in power_values.split(",")):
                        return "No Signal"
                    values = [float(v) for v in power_values.split(",") if
                              v.strip().replace('.', '', 1).replace('-', '', 1).isdigit()]
                    if all(-4 <= v <= 4 for v in values):
                        return "GOOD"
                    return "BAD"

                df.loc[signal_match.index, 'status_TX'] = evaluate_status(tx)
                df.loc[signal_match.index, 'status_RX'] = evaluate_status(rx)

        # Mark entries that haven't been matched
        df.loc[df['current_remote_host'].isnull() & df['current_remote_int'].isnull(), 'status'] = 'Линк отсутствует'

        # Save the final populated file
        df.to_excel(result_file, index=False)

        # Add borders to the Excel file
        self.add_borders_to_excel(result_file)

        return result_file




